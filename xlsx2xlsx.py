# coding: utf-8
from __future__ import unicode_literals
from io import StringIO
from openpyxl import load_workbook, Workbook

import tkFileDialog
import tkMessageBox
import ttk
from Tkinter import *

root = Tk()
root.geometry('500x300+400+300')
root.resizable(False, False)
treedata = []


def openfile(filename):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb.worksheets[0]
    temp = StringIO()

    for row in ws.rows:
        if row[0].value:
            temp.writelines(str(row[0].value) + '\n')
    return temp


def handle(fd):
    treedata = []
    fd.seek(0)
    
    title = False
    
    line = fd.readline().split(',')
    if line[0][:22] != 'RELEASE HEADER SECTION':
        tkMessageBox.showinfo('错误', '无法处理的xlsx')
        return
    while True:
        key, value, sched = [], [], []
        if line[0][:22] == 'RELEASE HEADER SECTION':
            key.extend(fd.readline().split(','))
            value.extend(fd.readline().split(','))

            line = fd.readline().split(',')
            key.extend(fd.readline().split(','))
            value.extend(fd.readline().split(','))

            line = fd.readline().split(',')
            key.extend(fd.readline().split(','))
            value.extend(fd.readline().split(','))

            line = fd.readline().split(',')
            key.extend(fd.readline().split(','))
            while True:
                line = fd.readline().split(',')
                if not line[0] or line[0][:22] == 'RELEASE HEADER SECTION':
                    break
                sched.append(line)

        if not title:
            treedata.append(key)
            title = True
        for i in range(len(sched)):
            row = []
            row.extend(value)
            row.extend(sched[i])
            if i > 0:
                row.append('DITTO')
            treedata.append(row)

        if not line[0]:
            break

    return treedata


def update_table(treedata):
    sy = Scrollbar(root)
    sy.pack(side=RIGHT, fill=Y)
    tree = ttk.Treeview(root, columns=treedata[0], show="headings", height=20)
    tree.configure(yscroll=sy.set)
    sy.config(command=tree.yview)

    for key in treedata[0]:
        tree.column(key, width=10, anchor='center')
        tree.heading(key, text=key)

    tree.pack()
    for row in treedata[1:]:
        tree.insert('', 1, values=row)

file_opt = {
    "defaultextension": ".xlsx",
    "filetypes": [('xlsx files', '.xlsx')],
    "parent": root
}


def askopenfilename():
    global treedata
    if treedata:
        tkMessageBox.showinfo("错误", '已打开过 需重启')
    else:
        filename = tkFileDialog.askopenfilename(**file_opt)
        if filename:
            fd = openfile(filename)
            treedata = handle(fd)
            if treedata:
                update_table(treedata)


def asksavefilename():
    global treedata
    if not treedata:
        tkMessageBox.showinfo("错误", '先打开需要处理的xlsx')
    else:
        filename = tkFileDialog.asksaveasfilename(**file_opt)
        if filename:
            wb = Workbook(write_only=True)
            ws = wb.create_sheet()
            for row in treedata:
                ws.append(row)
            wb.save(filename)


menubar = Menu(root)
filemenu = Menu(menubar, tearoff=0)
filemenu.add_command(label="打开xlsx", command=askopenfilename)
filemenu.add_command(label="导出xlsx", command=asksavefilename)
menubar.add_cascade(label="文件", menu=filemenu)
root.config(menu=menubar)

root.mainloop()
