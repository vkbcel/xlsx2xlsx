# coding: utf-8
from __future__ import unicode_literals
from io import StringIO
from openpyxl import load_workbook, Workbook


class XlsxHelper(object):

    def __init__(self):
        self.output = None
        self.treedata = None

    def openfile(self, filename):
        output = StringIO()
        wb = load_workbook(filename=filename, read_only=True)
        ws = wb.worksheets[0]

        for row in ws.rows:
            if row[0].value:
                output.writelines(str(row[0].value) + '\n')

        self.output = output

    def handle(self):
        treedata = []
        title = False

        self.output.seek(0)
    
        line = self.output.readline().split(',')
        if line[0][:22] != 'RELEASE HEADER SECTION':
            raise Exception('无法处理的xlsx')
        while True:
            key, value, sched = [], [], []
            if line[0][:22] == 'RELEASE HEADER SECTION':
                key.extend(self.output.readline().split(','))
                value.extend(self.output.readline().split(','))
    
                line = self.output.readline().split(',')
                key.extend(self.output.readline().split(','))
                value.extend(self.output.readline().split(','))
    
                line = self.output.readline().split(',')
                key.extend(self.output.readline().split(','))
                value.extend(self.output.readline().split(','))
    
                line = self.output.readline().split(',')
                key.extend(self.output.readline().split(','))
                while True:
                    line = self.output.readline().split(',')
                    if not line[0] or line[0][:22] == 'RELEASE HEADER SECTION':
                        break
                    sched.append(line)
    
            if not title:
                treedata.append(key)
                title = True
            for i in range(len(sched)):
                row = []
                row.extend(value)
                row.extend(sched[i])
                if i > 0:
                    row.append('DITTO')
                treedata.append(row)
    
            if not line[0]:
                break
    
        self.treedata = treedata

    def save(self, filename):
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        for row in self.treedata:
            ws.append(row)
        wb.save(filename)
